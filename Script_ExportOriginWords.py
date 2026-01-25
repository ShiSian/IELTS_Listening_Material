import os
from openpyxl import load_workbook

def export_to_origin_txt():
    excel_file = 'Jayden.xlsx'
    output_folder = 'OriginWords'
    template_name = 'Template'

    # 1. 检查文件和文件夹
    if not os.path.exists(excel_file):
        print(f"错误: 找不到文件 {excel_file}")
        return

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"已创建输出文件夹: {output_folder}")

    # 2. 加载工作簿 (data_only=True 确保读取的是公式计算后的值)
    wb = load_workbook(excel_file, data_only=True)

    for sheet_name in wb.sheetnames:
        # 跳过模板页
        if sheet_name == template_name:
            continue

        ws = wb[sheet_name]

        # 确定最后一行 (如果没有数据则跳过)
        max_row = ws.max_row
        if max_row < 3:
            print(f"Sheet '{sheet_name}' 没有有效数据行，跳过...")
            continue

        # 准备输出文件路径
        txt_file_path = os.path.join(output_folder, f"{sheet_name}.txt")

        # 提取第二列 (Column B)，从第三行到最后一行
        extracted_data = []
        for row_idx in range(3, max_row + 1):
            cell_value = ws.cell(row=row_idx, column=2).value

            # 只有当单元格不为空时才记录，或者保留空行（根据需求，通常建议保留以维持行号对应）
            if cell_value is not None:
                extracted_data.append(str(cell_value))
            else:
                extracted_data.append("")  # 如果希望空单元格也占一行

        # 3. 写入文本文件
        try:
            with open(txt_file_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(extracted_data))
            print(f"成功导出: {txt_file_path}")
        except Exception as e:
            print(f"导出 {sheet_name} 时出错: {e}")

    print("\n所有导出任务已完成！")

if __name__ == "__main__":
    export_to_origin_txt()