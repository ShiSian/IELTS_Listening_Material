import os
import re
import pandas as pd
from openpyxl import load_workbook

def try_int(s):
    """尝试将字符串转换为整数，用于排序"""
    try:
        return int(s)
    except:
        return s

def natural_key(string_):
    """将字符串拆分为数字和文字，生成用于自然排序的 key"""
    return [try_int(c) for c in re.split('([0-9]+)', string_)]

def process_csv_to_excel():
    excel_file = 'Jayden.xlsx'
    csv_folder = 'CSV'
    template_name = 'Template'

    if not os.path.exists(excel_file) or not os.path.exists(csv_folder):
        print("错误: 请检查 Excel 文件或 CSV 文件夹是否存在")
        return

    wb = load_workbook(excel_file)

    # 获取 CSV 文件列表并进行自然排序 (1, 2, 10...)
    csv_files = [f for f in os.listdir(csv_folder) if f.endswith('.csv')]
    csv_files.sort(key=natural_key)

    for file_name in csv_files:
        sheet_name = os.path.splitext(file_name)[0]
        csv_path = os.path.join(csv_folder, file_name)

        # 1. 确认并创建 Sheet
        if sheet_name not in wb.sheetnames:
            source = wb[template_name]
            target = wb.copy_worksheet(source)
            target.title = sheet_name
            print(f"已创建新 Sheet: {sheet_name}")

        ws = wb[sheet_name]

        # 2. 读取并写入数据
        df = pd.read_csv(csv_path, header=None)
        for i, row in df.iterrows():
            excel_row = i + 3
            ws.cell(row=excel_row, column=1, value=row[0])
            ws.cell(row=excel_row, column=2, value=row[1])
            ws.cell(row=excel_row, column=3, value=row[1])
            ws.cell(row=excel_row, column=4, value=row[2])

    # --- 关键修改：对 Excel 中的 Sheet 标签页进行物理排序 ---
    # 获取除了 Template 以外的所有 sheet 名，并按自然排序法排序
    all_sheets = wb.sheetnames
    other_sheets = [s for s in all_sheets if s != template_name]
    other_sheets.sort(key=natural_key)

    # 重新排列 wb._sheets 列表的顺序（Template 放在第一位，后面按数字顺序）
    ordered_names = [template_name] + other_sheets
    wb._sheets = [wb[name] for name in ordered_names]

    wb.save(excel_file)
    print("\n任务全部完成，Sheet 已按数字顺序排列！")

if __name__ == "__main__":
    process_csv_to_excel()