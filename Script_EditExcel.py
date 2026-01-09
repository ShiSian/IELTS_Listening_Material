import openpyxl
from openpyxl.utils import get_column_letter
import os

# 配置文件名
FILE_NAME = '王陆听力语料库.xlsx'
TARGET_WIDTH = 20.0

def main():
    # 检查文件是否存在
    if not os.path.exists(FILE_NAME):
        print(f"❌ 错误: 未在当前目录下找到文件 '{FILE_NAME}'")
        return

    try:
        print(f"正在加载 {FILE_NAME} ...")
        wb = openpyxl.load_workbook(FILE_NAME)

        print(f"正在处理 {len(wb.sheetnames)} 个Sheet...")

        for sheet in wb.worksheets:
            # 1. 设置工作表的“默认列宽”属性 (影响未被单独设置过的空白列)
            # 注意：Excel中默认宽度的单位换算比较特殊，但直接设为20通常能达到效果
            sheet.sheet_format.defaultColWidth = TARGET_WIDTH

            # 2. 强制覆盖现有有数据的列 (包括A-X列)
            # 如果只设默认值，某些已经被修改过宽度的列可能不会变，所以需要强制覆盖
            max_col = sheet.max_column
            # 如果是空表，至少处理到X列(24)，配合上一条脚本
            process_range = max(max_col, 24)

            for i in range(1, process_range + 1):
                col_letter = get_column_letter(i)
                sheet.column_dimensions[col_letter].width = TARGET_WIDTH

            print(f"  - {sheet.title}: 已设置默认宽度及前 {process_range} 列的宽度")

        # 保存文件
        print("正在保存文件...")
        wb.save(FILE_NAME)
        print(f"✅ 成功! 所有Sheet的列宽已设置为 {TARGET_WIDTH}。")

    except PermissionError:
        print(f"❌ 错误: 无法写入文件。请确保 '{FILE_NAME}' 没有在Excel中被打开。")
    except Exception as e:
        print(f"❌ 发生未知错误: {e}")

if __name__ == "__main__":
    main()