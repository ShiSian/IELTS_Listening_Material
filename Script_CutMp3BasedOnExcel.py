import glob

from pydub import AudioSegment, silence
import os
import openpyxl
from openpyxl.utils import column_index_from_string
from soupsieve.css_match import DIR_FLAGS

# 文件夹配置
DIR_ORIGIN_AUDIO = "OriginAudio"  # 存放原始 MP3 文件的文件夹
DIR_ORIGIN_WORDS = "OriginWords"    # 存放 Origin_*.txt 的文件夹
DIR_INTERMEDIATE = "Intermediate" # 存放 Keep_*.txt 的文件夹
DIR_OUTPUT       = "Output"           # 存放生成结果的文件夹

# ================= 导出指定Sheet的B列数据（仅导出未隐藏的行） ====================
def export_specific_sheets(excel_path, target_sheets):
    # 1. 检查文件是否存在
    if not os.path.exists(excel_path):
        print(f"错误：找不到文件 '{excel_path}'")
        return

    print(f"正在加载 Excel 文件: {excel_path} ... (文件较大时可能需要几秒)")

    # 2. 加载工作簿
    # data_only=True 确保读取的是计算后的数值而不是公式
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"无法打开Excel文件: {e}")
        return

    # 3. 遍历指定的 Sheet
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  警告：Excel中找不到名为 '{sheet_name}' 的Sheet，已跳过。")
            continue

        ws = wb[sheet_name]
        exported_data = []

        print(f"正在处理 Sheet: {sheet_name} ...")

        # 4. 遍历行 (从第3行开始)
        # openpyxl 的 max_row 会获取已使用的最大行数
        for row_idx in range(3, ws.max_row + 1):

            # --- 核心逻辑：检查行是否被隐藏 ---
            # ws.row_dimensions 存储了行的属性
            # 如果 row_idx 在 row_dimensions 中且 hidden 为 True，则跳过
            if row_idx in ws.row_dimensions and ws.row_dimensions[row_idx].hidden:
                # 这一行是隐藏的，跳过
                continue

            # --- 读取 B 列数据 (Column 2) ---
            cell_value = ws.cell(row=row_idx, column=2).value

            # 如果单元格不为空，则加入列表
            # 这里使用了 str() 确保数字也能被写入 txt
            if cell_value is not None:
                exported_data.append(str(cell_value))

        # 5. 写入 TXT 文件
        if exported_data:
            # 确保 Intermediate 文件夹存在
            os.makedirs("Intermediate", exist_ok=True)
            txt_filename = os.path.join("Intermediate", f"Keep_{sheet_name}.txt")
            try:
                with open(txt_filename, 'w', encoding='utf-8') as f:
                    f.write('\n'.join(exported_data))
                print(f"✅ 成功导出: {txt_filename} (共 {len(exported_data)} 行数据)")
            except Exception as e:
                print(f"❌ 写入文件 {txt_filename} 失败: {e}")
        else:
            print(f"⚠️  Sheet '{sheet_name}' 没有可导出的有效数据（A3之后为空或全为隐藏行）。")

    print("-" * 30)
    print("所有任务完成。")
     
def load_list(file_path):
    if not os.path.exists(file_path):
        return None
    with open(file_path, 'r', encoding='utf-8') as f:
        # 使用 strip() 去除换行符和首尾空格，并过滤空行
        return [line.strip() for line in f if line.strip()]

def process_single_unit(filename):
    # 1. 提取不带后缀的名称 (例如 "D1S1.mp3" -> "D1S1")
    base_name = os.path.splitext(filename)[0]

    # 使用全局变量构建路径
    mp3_path   = os.path.join(DIR_ORIGIN_AUDIO, f"{base_name}.mp3")
    full_txt   = os.path.join(DIR_ORIGIN_WORDS, f"{base_name}.txt")
    keep_txt   = os.path.join(DIR_INTERMEDIATE, f"Keep_{base_name}.txt")
    output_mp3 = os.path.join(DIR_OUTPUT,       f"Cutted_{base_name}.mp3")

    # 自动创建输出文件夹（如果不存在）
    os.makedirs(DIR_OUTPUT, exist_ok=True)

    if not os.path.exists(full_txt):
        print(f"'{full_txt}' 不存在，跳过音频切割")
        return

    if not os.path.exists(keep_txt):
        print(f"'{keep_txt}' 不存在，跳过音频切割")
        return

    print(f"\n>>> 正在处理: {mp3_path}")
    full_list = load_list(full_txt)
    keep_list = load_list(keep_txt)

    audio = AudioSegment.from_mp3(mp3_path)

    # 检测静音
    nonsilent_ranges = silence.detect_nonsilent(
        audio,
        min_silence_len=800, # 保持较大值防止切断词组
        silence_thresh=-45,
        seek_step=5
    )

    print(f"  - 单词总数: {len(full_list)}")
    print(f"  - 检测片段: {len(nonsilent_ranges)}")

    start_idx = 0
    word_map = {}
    word_ranges = nonsilent_ranges[start_idx:]

    for i in range(len(full_list)):
        if i >= len(word_ranges):
            print(f"  - ❌ 警告: 音频片段耗尽，单词 '{full_list[i]}' 及其后续单词无法匹配。")
            break

        current_word = full_list[i]
        start = word_ranges[i][0]

        # 确定切片终点：默认为下一个有效片段的起点（包含停顿）
        # 如果是列表最后一个单词，或者由于杂音导致 ranges 后面还有多余的，
        # 我们这里做一个保护：如果是full_list的最后一个词，直接取到文件末尾
        if i == len(full_list) - 1:
            chunk = audio[start:]
        else:
            # 正常情况：切到下一个对应的 ranges 起点
            # 注意：这里我们用 word_ranges[i+1] 是安全的，因为我们是按 full_list 遍历的
            if i + 1 < len(word_ranges):
                next_start = word_ranges[i+1][0]
                chunk = audio[start:next_start]
            else:
                # 极端情况：单词表还没完，但音频段用完了（前面情况C已经拦截，这里是双重保险）
                chunk = audio[start:]

        word_map[current_word] = chunk

    # 合成导出
    combined = AudioSegment.empty()
    success_count = 0
    missing_words = []

    for word in keep_list:
        if word in word_map:
            combined += word_map[word]
            success_count += 1
        else:
            missing_words.append(word)

    if missing_words:
        print(f"  - ⚠️ 以下单词未找到音频: {missing_words}")

    if success_count > 0:
        combined.export(output_mp3, format="mp3")
        print(f"  - ✅ 成功导出: {output_mp3}")

"""
@brief 根据规则隐藏指定Sheet的指定行
"""
def hide_completed_rows(file_name, target_sheets, target_columns):
    if not os.path.exists(file_name):
        return

    # ---------- 第一阶段：重置 (取消隐藏) ----------
    try:
        wb_reset = openpyxl.load_workbook(file_name)
        for sheet_name in target_sheets:
            if sheet_name in wb_reset.sheetnames:
                ws = wb_reset[sheet_name]
                # 取消隐藏所有行
                for r in range(1, ws.max_row + 1):
                    ws.row_dimensions[r].hidden = False
                # 清除筛选器
                ws.auto_filter.ref = None
        wb_reset.save(file_name)
        wb_reset.close()
    except PermissionError:
        print("[Error] 文件被占用，无法重置。请关闭 Excel 后重试。")
        return

    # ---------- 第二阶段：判定并隐藏 ----------
    try:
        # data_only=True 用于读取公式计算后的值（如果有缓存）
        wb_reader = openpyxl.load_workbook(file_name, data_only=True)
        # 普通模式用于写入隐藏属性（保留公式）
        wb_writer = openpyxl.load_workbook(file_name)
    except Exception as e:
        print(f"重新加载失败: {e}")
        return

    target_col_indices = [column_index_from_string(c) for c in target_columns]

    # === 修正后的等价计算函数 ===
    def calc_formula_equivalent(ws, row, formula_col_idx):
        """
        手动模拟 Excel 公式逻辑。
        
        假设逻辑：
        用户在 '公式列的左边一列' 输入内容。
        公式判断：IF(输入内容 == B列 OR 输入内容 == C列, "√", 错误提示)
        """
        def norm(v):
            return "" if v is None else str(v).strip()

        # 1. 获取标准答案 (B列:单词, C列:美音/解释, D列:中文)
        b = norm(ws[f"B{row}"].value)
        c = norm(ws[f"C{row}"].value)
        d = norm(ws[f"D{row}"].value)

        # 2. 获取用户输入值 (关键修改：读取公式列左侧的单元格)
        # 假设公式在 G 列 (index 7)，则输入在 F 列 (index 6)
        input_col_idx = formula_col_idx - 1
        user_input_val = ws.cell(row=row, column=input_col_idx).value
        user_input = norm(user_input_val)

        # 3. 逻辑判断
        if user_input == "":
            return ""

        # 如果输入内容等于 B列 或 C列，则视为正确
        if user_input.lower() == b.lower() or user_input.lower() == c.lower():
            return "√"

        # 否则返回错误提示格式
        return f"{b}|{c}>{d}"

    # 开始遍历处理
    for sheet_name in target_sheets:
        if sheet_name not in wb_reader.sheetnames:
            continue

        ws_read = wb_reader[sheet_name]
        ws_write = wb_writer[sheet_name]

        rows_hidden_count = 0
        max_row = ws_read.max_row

        # 从第3行开始（避开表头）
        for row_idx in range(3, max_row + 1):
            match_counter = 0

            for col_idx in target_col_indices:
                # 调用 Python 函数手动计算
                cell_value = calc_formula_equivalent(ws_read, row_idx, col_idx)

                # 格式化结果
                str_value = str(cell_value).strip() if cell_value is not None else ""

                if str_value == "√" or str_value == "":
                    match_counter += 1

            # 如果所有目标列都判定为 "√"，则隐藏该行
            if match_counter == len(target_col_indices):
                ws_write.row_dimensions[row_idx].hidden = True
                rows_hidden_count += 1

        print(f"  -> Sheet '{sheet_name}' 处理完毕，隐藏了 {rows_hidden_count} 行。")

    try:
        wb_writer.save(file_name)
        print(">>> 任务最终完成！")
    except PermissionError:
        print("[Error] 保存最终结果失败，请检查文件是否未关闭。")


 
        
    


def main():
    # 1、【这里可能需要修改】告诉程序操作哪个Excel文件
    file_to_handle = "Jayden.xlsx"
    
    # 王路语料库：内置sheet名称
    wanglu_chapter03 = ["3.2","3.3-1", "3.3-2", "3.3-3", "3.3-4", "3.3-5", "3.3-6", "3.3-7", "3.3-8", "3.3-9"]
    wanglu_chapter04 = ["4.2", "4.3-1", "4.3-2", "4.3-3", "4.4"]
    wanglu_chapter05 = ["5.2", "5.3-1", "5.3-2", "5.3-3", "5.3-4", "5.3-5", "5.3-6", "5.3-7", "5.3-8", "5.3-9", "5.3-10", "5.3-11", "5.3-12"]
    wanglu_chapter08 = ["8.2", "8.3-1", "8.3-2", "8.3-3", "8.3-4", "8.3-5", "8.4-1", "8.4-2", "8.4-3", "8.5", "8.6-1", "8.6-2", "8.6-3", "8.7-1", "8.7-2", "8.7-3", "8.8"]
    wanglu_chapter11 = ["11.1", "11.2", "11.3", "11.4"]
    wanglu_all_chapters = wanglu_chapter03 + wanglu_chapter04 + wanglu_chapter05 + wanglu_chapter08 + wanglu_chapter11
    # Jayden语料库：内置sheet名称
    jayden_d1  = ["D1S1", "D1S2"]
    jayden_d2  = ["D2S1", "D2S2"]
    jayden_d3  = ["D3S1", "D3S2"]
    jayden_d4  = ["D4S1", "D4S2"]
    jayden_d5  = ["D5S1", "D5S2"]
    jayden_d6  = ["D6S1", "D6S2"]
    jayden_d7  = ["D7S1", "D7S2"]
    jayden_d8  = ["D8S1", "D8S2"]
    jayden_d9  = ["D9S1", "D9S2"]
    jayden_d10 = ["D10S1", "D10S2"]
    jayden_d11 = ["D11S1", "D11S2"]
    jayden_d12 = ["D12S1", "D12S2"]
    jayden_d13 = ["D13S1", "D13S2"]
    jayden_d14 = ["D14S1", "D14S2"]
    jayden_d15 = ["D15S1", "D15S2"]
    jayden_d16 = ["D16S1", "D16S2"]
    jayden_d17 = ["D17S1", "D17S2"]
    jayden_d18 = ["D18S1", "D18S2"]
    jayden_d19 = ["D19S1", "D19S2"]
    jayden_d20 = ["D20S1", "D20S2"]
    jayden_d21 = ["D21S1", "D21S2"]    
    # 如果你需要一次性处理所有天数，可以使用这个合并列表
    jayden_all_sheets = (
            jayden_d1 + jayden_d2 + jayden_d3 + jayden_d4 + jayden_d5 + jayden_d6 +
            jayden_d7 + jayden_d8 + jayden_d9 + jayden_d10 + jayden_d11 + jayden_d12 +
            jayden_d13 + jayden_d14 + jayden_d15 + jayden_d16 + jayden_d17 + jayden_d18 +
            jayden_d19 + jayden_d20 + jayden_d21
    )
    
    
    # 【这里可能需要修改】告诉程序需要处理哪些单元表
    sheets_to_handle  = jayden_d1
    # 【这里可能需要修改】告诉程序基于哪几列的值来判断是否需要隐藏对应行（比如FGH表示如果在一行中FGH列的值都是√或者为空<表示本次不需要听写>，则隐藏该行）
    columns_to_handle = ["F", "H"]
    
    # 2、行隐藏操作
    hide_completed_rows(file_to_handle, sheets_to_handle, columns_to_handle)
    
    # 3、从Excel中导出需要keep的单词列表
    export_specific_sheets(file_to_handle, sheets_to_handle)

    # 4、基于 keep 单词列表切割 mp3
    # 检查原始音频文件夹是否存在
    if not os.path.exists(DIR_ORIGIN_AUDIO):
        print(f"❌ 错误：找不到原始音频文件夹 '{DIR_ORIGIN_AUDIO}'")
        return
    # 获取文件夹内所有文件
    all_files = os.listdir(DIR_ORIGIN_AUDIO)
    # 过滤出需要处理的 MP3
    for filename in all_files:
        # 只处理 .mp3 后缀
        if not filename.lower().endswith('.mp3'):
            continue
        file_base_name = os.path.splitext(filename)[0]
        # 匹配当前的 Sheet 名称（确保只处理本次指定的章节）
        if file_base_name not in sheets_to_handle:
            continue
        # 调用处理函数
        process_single_unit(filename)

if __name__ == "__main__":
    main()